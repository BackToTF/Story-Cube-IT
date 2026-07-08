from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

project = Path(r"C:\Users\ciro.andreano\Desktop\Story Cube - I&D")
share = Path(r"C:\Users\ciro.andreano\OneDrive - Avanade\Story Cube - I&D - Documenti\Story Cube - I&D")
out_path = project / "docs" / "adattamento gioco" / "StoryCube_Mappatura_Archetipi.xlsx"
out_path.parent.mkdir(parents=True, exist_ok=True)

archetypes = ["Creativo", "Analitico", "Empatico", "Innovatore", "Connector"]

cards = [
    ("P01", "Personaggio", "Un bambino curioso.", "Creativo", "Innovatore"),
    ("P02", "Personaggio", "Una bambina coraggiosa.", "Innovatore", "Empatico"),
    ("P03", "Personaggio", "Un inventore gentile.", "Innovatore", "Analitico"),
    ("P04", "Personaggio", "Un'amica che ascolta tutti.", "Empatico", "Connector"),
    ("P05", "Personaggio", "Un fratello maggiore responsabile.", "Analitico", "Empatico"),
    ("P06", "Personaggio", "Una guida del villaggio.", "Connector", "Analitico"),
    ("L01", "Luogo", "Nel parco del quartiere.", "Empatico", "Creativo"),
    ("L02", "Luogo", "In una biblioteca colorata.", "Analitico", "Creativo"),
    ("L03", "Luogo", "In una scuola di magia.", "Creativo", "Innovatore"),
    ("L04", "Luogo", "Nel bosco vicino casa.", "Innovatore", "Empatico"),
    ("L05", "Luogo", "In una piazza piena di giochi.", "Connector", "Empatico"),
    ("L06", "Luogo", "In un laboratorio creativo.", "Analitico", "Innovatore"),
    ("O01", "Oggetto", "Una mappa con indizi.", "Analitico", "Creativo"),
    ("O02", "Oggetto", "Una chiave dorata.", "Innovatore", "Creativo"),
    ("O03", "Oggetto", "Uno zaino con strumenti utili.", "Analitico", "Connector"),
    ("O04", "Oggetto", "Una lanterna luminosa.", "Empatico", "Innovatore"),
    ("O05", "Oggetto", "Un libro di idee.", "Creativo", "Analitico"),
    ("O06", "Oggetto", "Una scatola dei messaggi.", "Connector", "Empatico"),
    ("R01", "Problema", "La strada giusta non si trova.", "Analitico", "Connector"),
    ("R02", "Problema", "Il gruppo litiga su cosa fare.", "Empatico", "Connector"),
    ("R03", "Problema", "Un amico si sente escluso.", "Empatico", "Connector"),
    ("R04", "Problema", "Un oggetto importante sparisce.", "Analitico", "Innovatore"),
    ("R05", "Problema", "Sta arrivando la pioggia forte.", "Innovatore", "Analitico"),
    ("R06", "Problema", "Il tempo per finire sta per scadere.", "Analitico", "Creativo"),
    ("F01", "Finale", "Il gruppo risolve insieme.", "Connector", "Empatico"),
    ("F02", "Finale", "Ogni personaggio aiuta con un talento.", "Connector", "Creativo"),
    ("F03", "Finale", "Si trova una soluzione nuova.", "Innovatore", "Creativo"),
    ("F04", "Finale", "L'amico escluso torna al centro.", "Empatico", "Connector"),
    ("F05", "Finale", "Il problema diventa una lezione.", "Analitico", "Empatico"),
    ("F06", "Finale", "La storia finisce con una festa.", "Creativo", "Empatico"),
]

wb = Workbook()
ws_catalog = wb.active
ws_catalog.title = "Catalogo_Risposte"
ws_input = wb.create_sheet("Input_Semplice")
ws_score = wb.create_sheet("Score_Archetipi")
ws_notes = wb.create_sheet("Istruzioni")

header_fill = PatternFill("solid", fgColor="F56A00")
header_font = Font(color="FFFFFF", bold=True)
edit_fill = PatternFill("solid", fgColor="FFF1E8")

catalog_headers = ["ID_Risposta", "Fase", "Risposta_Semplice", "Archetipo_Primario", "Archetipo_Secondario", "Da_Revisionare"]
ws_catalog.append(catalog_headers)
for c in range(1, len(catalog_headers) + 1):
    cell = ws_catalog.cell(1, c)
    cell.fill = header_fill
    cell.font = header_font
for row in cards:
    ws_catalog.append([row[0], row[1], row[2], row[3], row[4], "SI"])
for col, width in {"A":14, "B":14, "C":55, "D":22, "E":22, "F":16}.items():
    ws_catalog.column_dimensions[col].width = width
arch_list = '"' + ','.join(archetypes) + '"'
dv_arch = DataValidation(type="list", formula1=arch_list, allow_blank=True)
ws_catalog.add_data_validation(dv_arch)
dv_arch.add("D2:E200")

input_headers = [
    "Data", "Sessione", "Player", "Personaggio_ID", "Luogo_ID", "Oggetto_ID", "Problema_ID", "Finale_ID",
    "Personaggio", "Luogo", "Oggetto", "Problema", "Finale",
    "A1_P", "A2_P", "A1_L", "A2_L", "A1_O", "A2_O", "A1_R", "A2_R", "A1_F", "A2_F",
]
ws_input.append(input_headers)
for c in range(1, len(input_headers) + 1):
    cell = ws_input.cell(1, c)
    cell.fill = header_fill
    cell.font = header_font

for r in range(2, 202):
    ws_input.cell(r, 9).value = f'=IFERROR(INDEX(Catalogo_Risposte!$C:$C, MATCH($D{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 10).value = f'=IFERROR(INDEX(Catalogo_Risposte!$C:$C, MATCH($E{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 11).value = f'=IFERROR(INDEX(Catalogo_Risposte!$C:$C, MATCH($F{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 12).value = f'=IFERROR(INDEX(Catalogo_Risposte!$C:$C, MATCH($G{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 13).value = f'=IFERROR(INDEX(Catalogo_Risposte!$C:$C, MATCH($H{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 14).value = f'=IFERROR(INDEX(Catalogo_Risposte!$D:$D, MATCH($D{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 15).value = f'=IFERROR(INDEX(Catalogo_Risposte!$E:$E, MATCH($D{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 16).value = f'=IFERROR(INDEX(Catalogo_Risposte!$D:$D, MATCH($E{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 17).value = f'=IFERROR(INDEX(Catalogo_Risposte!$E:$E, MATCH($E{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 18).value = f'=IFERROR(INDEX(Catalogo_Risposte!$D:$D, MATCH($F{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 19).value = f'=IFERROR(INDEX(Catalogo_Risposte!$E:$E, MATCH($F{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 20).value = f'=IFERROR(INDEX(Catalogo_Risposte!$D:$D, MATCH($G{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 21).value = f'=IFERROR(INDEX(Catalogo_Risposte!$E:$E, MATCH($G{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 22).value = f'=IFERROR(INDEX(Catalogo_Risposte!$D:$D, MATCH($H{r}, Catalogo_Risposte!$A:$A, 0)), "")'
    ws_input.cell(r, 23).value = f'=IFERROR(INDEX(Catalogo_Risposte!$E:$E, MATCH($H{r}, Catalogo_Risposte!$A:$A, 0)), "")'

for col, width in {
    "A":12, "B":12, "C":12, "D":14, "E":12, "F":12, "G":13, "H":11,
    "I":28, "J":28, "K":28, "L":28, "M":28,
    "N":10, "O":10, "P":10, "Q":10, "R":10, "S":10, "T":10, "U":10, "V":10, "W":10,
}.items():
    ws_input.column_dimensions[col].width = width

dv_p = DataValidation(type="list", formula1='"P01,P02,P03,P04,P05,P06"', allow_blank=True)
dv_l = DataValidation(type="list", formula1='"L01,L02,L03,L04,L05,L06"', allow_blank=True)
dv_o = DataValidation(type="list", formula1='"O01,O02,O03,O04,O05,O06"', allow_blank=True)
dv_r = DataValidation(type="list", formula1='"R01,R02,R03,R04,R05,R06"', allow_blank=True)
dv_f = DataValidation(type="list", formula1='"F01,F02,F03,F04,F05,F06"', allow_blank=True)
for dv, rng in [(dv_p, "D2:D201"), (dv_l, "E2:E201"), (dv_o, "F2:F201"), (dv_r, "G2:G201"), (dv_f, "H2:H201")]:
    ws_input.add_data_validation(dv)
    dv.add(rng)

score_headers = ["Player"] + archetypes + ["Max", "Profilo_Risultante"]
ws_score.append(score_headers)
for c in range(1, len(score_headers) + 1):
    cell = ws_score.cell(1, c)
    cell.fill = header_fill
    cell.font = header_font

for r in range(2, 42):
    ws_score.cell(r, 2).value = f'=COUNTIFS(Input_Semplice!$C:$C,$A{r},Input_Semplice!$N:$W,B$1)'
    ws_score.cell(r, 3).value = f'=COUNTIFS(Input_Semplice!$C:$C,$A{r},Input_Semplice!$N:$W,C$1)'
    ws_score.cell(r, 4).value = f'=COUNTIFS(Input_Semplice!$C:$C,$A{r},Input_Semplice!$N:$W,D$1)'
    ws_score.cell(r, 5).value = f'=COUNTIFS(Input_Semplice!$C:$C,$A{r},Input_Semplice!$N:$W,E$1)'
    ws_score.cell(r, 6).value = f'=COUNTIFS(Input_Semplice!$C:$C,$A{r},Input_Semplice!$N:$W,F$1)'
    ws_score.cell(r, 7).value = f'=MAX(B{r}:F{r})'
    ws_score.cell(r, 8).value = f'=IF($A{r}="","",TEXTJOIN(" + ",TRUE,IF(B{r}:F{r}=G{r},$B$1:$F$1,"")))'

for col, width in {"A":14, "B":14, "C":14, "D":14, "E":14, "F":14, "G":10, "H":34}.items():
    ws_score.column_dimensions[col].width = width

ws_notes["A1"] = "Istruzioni rapide (versione semplificata)"
ws_notes["A1"].font = Font(bold=True, size=14)
notes = [
    "1) Ogni riga di Input_Semplice e una mini-storia del player.",
    "2) Inserisci Data, Sessione, Player.",
    "3) Scegli 1 carta per fase: Personaggio, Luogo, Oggetto, Problema, Finale.",
    "4) I testi delle carte e gli archetipi si compilano da soli.",
    "5) In Score_Archetipi scrivi i nomi player in colonna A.",
    "6) Profilo_Risultante mostra archetipo dominante o mix in caso di parita.",
    "Nota: tutte le carte hanno testo semplice e combinabile logicamente.",
]
for i, text in enumerate(notes, start=3):
    ws_notes[f"A{i}"] = text
ws_notes.column_dimensions["A"].width = 120

for ws in [ws_catalog, ws_input, ws_score]:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

for ws, rng in [
    (ws_catalog, "D2:F200"),
    (ws_input, "A2:H201"),
    (ws_score, "A2:A41"),
]:
    for row in ws[rng]:
        for cell in row:
            cell.fill = edit_fill

wb.save(out_path)

sp_out = share / "docs" / "adattamento gioco" / out_path.name
sp_out.parent.mkdir(parents=True, exist_ok=True)
sp_out.write_bytes(out_path.read_bytes())

print(f"CREATED={out_path}")
print(f"SYNCED={sp_out}")
print(f"CARDS={len(cards)}")
